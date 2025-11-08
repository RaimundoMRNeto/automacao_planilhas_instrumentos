# ============================================================
# ETL SICONV – Modo de Teste (com pagamento, tributo e OBTV)
# - Sem verificação de data_carga e sem state
# - CONVENIOS: mantém TODAS as colunas originais + ordem NR_CONVENIO→texto→data→numérico→outros
# - PAGAMENTO, PAGAMENTO_TRIBUTO, OBTV_CONVENENTE: mantém TODAS as colunas, ordenadas alfabeticamente
# - XLSX com: convenios, pagamento, pagamento_tributo, obtv_convenente, dicionario_variaveis, info_execucao
# - Valores monetários como número (R$ pt-BR no Excel)
# - Backups por data/hora de EXECUÇÃO (YYYY/MM/DD + timestamp)
# ============================================================

from __future__ import annotations
import io, sys, zipfile, shutil
from datetime import datetime, timezone
from pathlib import Path
from typing import List, Dict, Tuple, Optional

import pandas as pd
import requests
import yaml
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from email.utils import parsedate_to_datetime

try:
    from zoneinfo import ZoneInfo
    Z_SP = ZoneInfo("America/Sao_Paulo"); Z_UTC = ZoneInfo("UTC")
except Exception:
    Z_SP = None; Z_UTC = None

# -----------------------
# URLs do Repositório DETRU
# -----------------------
URL_BASE = "https://repositorio.dados.gov.br/seges/detru/"
URL_DATA_CARGA = URL_BASE + "data_carga_siconv.csv.zip"
URL_CONVENIO   = URL_BASE + "siconv_convenio.csv.zip"
URL_PAGAMENTO  = URL_BASE + "siconv_pagamento.csv.zip"
URL_PAG_TRIB   = URL_BASE + "siconv_pagamento_tributo.csv.zip"
URL_OBTV_CONV  = URL_BASE + "siconv_obtv_convenente.csv.zip"

# -----------------------
# Pastas/arquivos locais
# -----------------------
DIR_OUT   = Path("out")
DIR_CACHE = Path("data/cache")
PATH_IDS        = Path("config/ids_convenio.txt")
PATH_COLUMNSCFG = Path("config/columns_convenio.yaml")

SHEET_CONV = "convenios"
SHEET_PAG  = "pagamento"
SHEET_TRIB = "pagamento_tributo"
SHEET_OBTV = "obtv_convenente"
DICT_SHEET = "dicionario_variaveis"
INFO_SHEET = "info_execucao"

# Colunas monetárias conhecidas
MONEY_COLS = [
    # convenio
    "VL_GLOBAL_CONV","VL_REPASSE_CONV","VL_CONTRAPARTIDA_CONV","VL_EMPENHADO_CONV","VL_DESEMBOLSADO_CONV",
    "VL_SALDO_CONTA","VL_SALDO_REMAN_TESOURO","VL_SALDO_REMAN_CONVENENTE","VL_RENDIMENTO_APLICACAO",
    "VL_INGRESSO_CONTRAPARTIDA","VALOR_GLOBAL_ORIGINAL_CONV",
    # pagamento / tributo / obtv
    "VL_PAGO","VL_PAG_TRIBUTOS","VL_PAGO_OBTV_CONV",
]

def log(msg: str) -> None:
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{ts}] {msg}")

def ensure_dirs() -> None:
    for p in [DIR_OUT, DIR_CACHE]:
        p.mkdir(parents=True, exist_ok=True)

def read_yaml_config(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}

def read_ids_convenio(path: Path) -> List[str]:
    if not path.exists():
        log(f"⚠️  {path} não encontrado — lista de IDs vazia.")
        return []
    with path.open("r", encoding="utf-8") as f:
        return [line.strip() for line in f if line.strip()]

def download_bytes(url: str) -> Tuple[bytes, Optional[str]]:
    log(f"↓ Baixando: {url}")
    r = requests.get(url, timeout=60)
    r.raise_for_status()
    return r.content, r.headers.get("Last-Modified")

def unzip_first_csv(zip_bytes: bytes) -> pd.DataFrame:
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as z:
        csv_names = [n for n in z.namelist() if n.lower().endswith(".csv")]
        if not csv_names:
            raise RuntimeError("Nenhum CSV encontrado no ZIP.")
        with z.open(csv_names[0]) as f:
            df = pd.read_csv(f, sep=";", encoding="utf-8", dtype=str, low_memory=False)
    df.columns = [c.strip().upper().replace(" ", "_") for c in df.columns]
    return df

def http_last_modified_to_strings(lm: Optional[str]) -> Tuple[Optional[str], Optional[str]]:
    if not lm:
        return None, None
    try:
        dt = parsedate_to_datetime(lm)
        utc_iso = dt.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        sp_str = dt.astimezone(Z_SP).strftime("%Y-%m-%d %H:%M:%S") if Z_SP else None
        return utc_iso, sp_str
    except Exception:
        return None, None

def classify_columns(df: pd.DataFrame, tipos_cfg: dict) -> Dict[str, List[str]]:
    texto_cfg = set([c.upper() for c in tipos_cfg.get("texto", [])])
    data_cfg  = set([c.upper() for c in tipos_cfg.get("data", [])])
    num_cfg   = set([c.upper() for c in tipos_cfg.get("numerico", [])])
    texto, datas, numeros, outros = [], [], [], []
    for c in df.columns:
        if c == "NR_CONVENIO":
            continue
        cu = c.upper()
        if cu in texto_cfg:
            texto.append(c)
        elif cu in data_cfg:
            datas.append(c)
        elif cu in num_cfg or cu in MONEY_COLS:
            numeros.append(c)
        else:
            outros.append(c)
    return {"texto": texto, "data": datas, "numerico": numeros, "outros": outros}

def order_columns_convenio(df: pd.DataFrame, cfg: dict) -> List[str]:
    grupos = classify_columns(df, cfg.get("tipos_colunas", {}))
    final: List[str] = []
    if "NR_CONVENIO" in df.columns:
        final.append("NR_CONVENIO")
    for grupo in ["texto", "data", "numerico", "outros"]:
        final.extend([c for c in df.columns if c in grupos.get(grupo, [])])
    seen, ordered = set(), []
    for c in final:
        if c not in seen:
            ordered.append(c); seen.add(c)
    for c in df.columns:
        if c not in seen:
            ordered.append(c); seen.add(c)
    return ordered

def order_columns_alpha(df: pd.DataFrame) -> List[str]:
    return sorted(df.columns.tolist())

def parse_monetary_series(s: pd.Series) -> pd.Series:
    if s.dtype.kind != "O":
        return pd.to_numeric(s, errors="coerce")
    x = s.astype(str).str.strip()
    x = x.str.replace("R$", "", regex=False).str.replace(" ", "", regex=False)
    def to_float(v: str):
        if v in ("", "nan", "None", "NULL"):
            return None
        try:
            return float(v.replace(".", "").replace(",", "."))  # 1.234,56
        except Exception:
            try:
                return float(v.replace(",", ""))                  # 1,234.56
            except Exception:
                return None
    return x.map(to_float)

# === dicionário (descrições) ===
DESC_MAP = {
    # convenio.html
    "NR_CONVENIO": "Número gerado pelo Siconv (faixa reservada 700000–999999).",
    "ID_PROPOSTA": "Código sequencial do sistema para uma Proposta.",
    "DIA": "Dia em que o convênio foi assinado.",
    "MES": "Mês em que o convênio foi assinado.",
    "ANO": "Ano de assinatura do convênio.",
    "DIA_ASSIN_CONV": "Data de assinatura do convênio.",
    "SIT_CONVENIO": "Situação atual do convênio.",
    "SUBSITUACAO_CONV": "Sub-situação atual do convênio.",
    "SITUACAO_PUBLICACAO": "Situação da publicação do instrumento.",
    "INSTRUMENTO_ATIVO": "Indica se o instrumento não foi finalizado (SIM/NÃO).",
    "IND_OPERA_OBTV": "Indica operação com OBTV (SIM/NÃO).",
    "NR_PROCESSO": "Número interno do processo físico.",
    "UG_EMITENTE": "Número da Unidade Gestora.",
    "DIA_PUBL_CONV": "Data da publicação do convênio.",
    "DIA_INIC_VIGENC_CONV": "Data de início da vigência.",
    "DIA_FIM_VIGENC_CONV": "Data de fim da vigência.",
    "DIA_FIM_VIGENC_ORIGINAL_CONV": "Fim de vigência original (sem TAs/Prorrogas).",
    "DIAS_PREST_CONTAS": "Prazo (dias) para prestação de contas.",
    "DIA_LIMITE_PREST_CONTAS": "Data limite para prestação de contas.",
    "DATA_SUSPENSIVA": "Previsão de resolução da cláusula suspensiva.",
    "DATA_RETIRADA_SUSPENSIVA": "Data de retirada da cláusula suspensiva.",
    "DIAS_CLAUSULA_SUSPENSIVA": "Dias entre previsão e assinatura.",
    "SITUACAO_CONTRATACAO": "Situação da contratação.",
    "IND_ASSINADO": "Indicador de convênio assinado (SIM/NÃO).",
    "MOTIVO_SUSPENSAO": "Motivo de suspensão (cláusula suspensiva).",
    "IND_FOTO": "Indicador se possui foto (SIM/NÃO).",
    "QTDE_CONVENIOS": "Quantidade de instrumentos assinados.",
    "QTD_TA": "Quantidade de termos aditivos.",
    "QTD_PRORROGA": "Quantidade de prorrogações de ofício.",
    "VL_GLOBAL_CONV": "Valor global (repasse + contrapartida).",
    "VL_REPASSE_CONV": "Valor total do repasse da União.",
    "VL_CONTRAPARTIDA_CONV": "Valor total da contrapartida.",
    "VL_EMPENHADO_CONV": "Valor total empenhado.",
    "VL_DESEMBOLSADO_CONV": "Valor total desembolsado.",
    "VL_SALDO_REMAN_TESOURO": "Valores devolvidos ao Tesouro ao término.",
    "VL_SALDO_REMAN_CONVENENTE": "Valores devolvidos ao Convenente ao término.",
    "VL_RENDIMENTO_APLICACAO": "Rendimentos de aplicação financeira.",
    "VL_INGRESSO_CONTRAPARTIDA": "Ingressos de contrapartida.",
    "VL_SALDO_CONTA": "Saldo em conta (estimado).",
    "VALOR_GLOBAL_ORIGINAL_CONV": "Valor global original do instrumento.",
    # pagamento.html
    "NR_MOV_FIN": "Número identificador da movimentação financeira.",
    "IDENTIF_FORNECEDOR": "CNPJ/CPF do fornecedor.",
    "NOME_FORNECEDOR": "Nome do fornecedor.",
    "TP_MOV_FINANCEIRA": "Tipo da movimentação (ex.: Pagamento a favorecido / com OBTV).",
    "DATA_PAG": "Data do pagamento.",
    "NR_DL": "Número do documento de liquidação.",
    "DESC_DL": "Descrição do DL (ex.: NOTA FISCAL, FATURA, etc.).",
    "VL_PAGO": "Valor do pagamento.",
    "ID_DL": "Identificador do documento de liquidação.",
    "DATA_EMISSAO_DL": "Data de emissão do documento de liquidação.",
    # pagamento_tributo.html
    "DATA_TRIBUTO": "Data de pagamento do tributo.",
    "VL_PAG_TRIBUTOS": "Valor do tributo.",
    # obtv_convenente.html
    "IDENTIF_FAVORECIDO_OBTV_CONV": "CNPJ/CPF do favorecido recebedor (OBTV).",
    "NM_FAVORECIDO_OBTV_CONV": "Nome do favorecido recebedor (OBTV).",
    "TP_AQUISICAO": "Tipo de aquisição.",
    "VL_PAGO_OBTV_CONV": "Valor pago ao favorecido (OBTV).",
}

def build_dictionary_df(dfs_by_sheet: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows = []
    for sheet, df in dfs_by_sheet.items():
        for col in df.columns:
            rows.append({"COLUNA": col, "DESCRICAO": DESC_MAP.get(col, "")})
    dict_df = pd.DataFrame(rows, columns=["COLUNA", "DESCRICAO"])
    dict_df = dict_df.drop_duplicates(subset=["COLUNA"], keep="first").sort_values("COLUNA").reset_index(drop=True)
    return dict_df

def build_info_sheet_df(
    run_dt_local: datetime,
    data_carga_raw: str,
    convs_encontrados: List[str],
    ids_total: int,
    extras: Dict[str, str],
    stats: Dict[str, int],
) -> pd.DataFrame:
    info_rows = [
        ("data_execucao_local", run_dt_local.strftime("%Y-%m-%d %H:%M:%S")),
        ("data_carga_publicada_csv", data_carga_raw or ""),
        ("fonte_convenio_zip", URL_CONVENIO),
        ("fonte_data_carga_zip", URL_DATA_CARGA),
        ("fonte_pagamento_zip", URL_PAGAMENTO),
        ("fonte_pagamento_tributo_zip", URL_PAG_TRIB),
        ("fonte_obtv_convenente_zip", URL_OBTV_CONV),
        ("instrumentos_encontrados", str(len(convs_encontrados))),
        ("instrumentos_selecionados", str(ids_total)),
        ("lista_encontrados", ", ".join(sorted(convs_encontrados))),
        ("registros_pagamento", str(stats.get("pagamento", 0))),
        ("registros_pagamento_tributo", str(stats.get("pagamento_tributo", 0))),
        ("registros_obtv_convenente", str(stats.get("obtv_convenente", 0))),
    ]
    # incluir (opcionalmente) timestamps HTTP Last-Modified
    for k, v in extras.items():
        info_rows.append((k, v or ""))
    return pd.DataFrame(info_rows, columns=["chave", "valor"])

def apply_currency_format(ws_name: str, wb, money_cols: List[str]):
    if ws_name not in wb.sheetnames:
        return
    ws = wb[ws_name]
    col_idx = {cell.value: i + 1 for i, cell in enumerate(ws[1]) if cell.value}
    br_currency = '[$R$-pt-BR] #,##0.00'
    for colname in money_cols:
        if colname in col_idx:
            j = col_idx[colname]
            letter = get_column_letter(j)
            for cell in ws[letter][1:]:
                cell.number_format = br_currency
    # auto largura simples (amostra até 100 linhas)
    for j in range(1, wb[ws_name].max_column + 1):
        letter = get_column_letter(j)
        max_len = 0
        for cell in wb[ws_name][letter][:100]:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        wb[ws_name].column_dimensions[letter].width = min(60, max(10, max_len * 0.9))

def write_excel_with_formats(dfs: Dict[str, pd.DataFrame], dict_df: pd.DataFrame, info_df: pd.DataFrame, path_xlsx: Path) -> None:
    path_xlsx.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path_xlsx, engine="openpyxl") as writer:
        for sheet, df in dfs.items():
            df.to_excel(writer, sheet_name=sheet, index=False)
        dict_df.to_excel(writer, sheet_name=DICT_SHEET, index=False)
        info_df.to_excel(writer, sheet_name=INFO_SHEET, index=False)

    wb = load_workbook(path_xlsx)
    # formatos monetários nas abas correspondentes
    apply_currency_format(SHEET_CONV, wb, MONEY_COLS)
    apply_currency_format(SHEET_PAG, wb, ["VL_PAGO"])
    apply_currency_format(SHEET_TRIB, wb, ["VL_PAG_TRIBUTOS"])
    apply_currency_format(SHEET_OBTV, wb, ["VL_PAGO_OBTV_CONV"])
    wb.save(path_xlsx)

# -----------------------
# Execução (SEM STATE / SEM BLOQUEIO DIÁRIO)
# -----------------------
def main() -> int:
    ensure_dirs()
    run_dt_local = datetime.now()

    # configs e IDs
    cfg = read_yaml_config(PATH_COLUMNSCFG)
    ids_lista = read_ids_convenio(PATH_IDS)

    # data_carga (apenas info_execucao)
    try:
        carga_bytes, lm_carga = download_bytes(URL_DATA_CARGA)
        df_carga = unzip_first_csv(carga_bytes)
        data_carga_raw = None
        for col in df_carga.columns:
            vals = df_carga[col].dropna()
            if not vals.empty:
                data_carga_raw = str(vals.iloc[0]).strip()
                if data_carga_raw: break
        if not data_carga_raw:
            data_carga_raw = run_dt_local.strftime("%Y-%m-%d")
        lm_carga_utc, lm_carga_sp = http_last_modified_to_strings(lm_carga)
    except Exception:
        data_carga_raw = run_dt_local.strftime("%Y-%m-%d")
        lm_carga_utc, lm_carga_sp = None, None

    # --- CONVENIO
    conv_bytes, lm_conv = download_bytes(URL_CONVENIO)
    df_conv_all = unzip_first_csv(conv_bytes)
    if "NR_CONVENIO" not in df_conv_all.columns:
        log("❌ Coluna 'NR_CONVENIO' não encontrada em CONVENIO.")
        return 1
    df_conv = df_conv_all[df_conv_all["NR_CONVENIO"].isin(ids_lista)].copy()
    log(f"🎯 Convênios encontrados: {len(df_conv)}")
    # moedas -> float
    for col in MONEY_COLS:
        if col in df_conv.columns:
            df_conv[col] = parse_monetary_series(df_conv[col])
    # ordenar colunas (regra preservada para convenio)
    cols_conv = order_columns_convenio(df_conv, cfg)
    df_conv = df_conv.loc[:, cols_conv]

    conv_encontrados = df_conv["NR_CONVENIO"].astype(str).unique().tolist()

    # --- PAGAMENTO
    pag_bytes, lm_pag = download_bytes(URL_PAGAMENTO)
    df_pag_all = unzip_first_csv(pag_bytes)
    if "NR_CONVENIO" in df_pag_all.columns:
        df_pag = df_pag_all[df_pag_all["NR_CONVENIO"].isin(ids_lista)].copy()
    else:
        df_pag = df_pag_all.iloc[0:0].copy()
    # parse valores
    for col in ["VL_PAGO"]:
        if col in df_pag.columns:
            df_pag[col] = parse_monetary_series(df_pag[col])
    # ordenar alfabético
    df_pag = df_pag.loc[:, order_columns_alpha(df_pag)] if not df_pag.empty else df_pag

    # --- PAGAMENTO_TRIBUTO
    trib_bytes, lm_trib = download_bytes(URL_PAG_TRIB)
    df_trib_all = unzip_first_csv(trib_bytes)
    if "NR_CONVENIO" in df_trib_all.columns:
        df_trib = df_trib_all[df_trib_all["NR_CONVENIO"].isin(ids_lista)].copy()
    else:
        df_trib = df_trib_all.iloc[0:0].copy()
    for col in ["VL_PAG_TRIBUTOS"]:
        if col in df_trib.columns:
            df_trib[col] = parse_monetary_series(df_trib[col])
    df_trib = df_trib.loc[:, order_columns_alpha(df_trib)] if not df_trib.empty else df_trib

    # --- OBTV_CONVENENTE (filtra por NR_MOV_FIN existente em pagamentos filtrados)
    obtv_bytes, lm_obtv = download_bytes(URL_OBTV_CONV)
    df_obtv_all = unzip_first_csv(obtv_bytes)
    if "NR_MOV_FIN" in df_obtv_all.columns and not df_pag.empty and "NR_MOV_FIN" in df_pag.columns:
        df_obtv = df_obtv_all[df_obtv_all["NR_MOV_FIN"].isin(df_pag["NR_MOV_FIN"])].copy()
    else:
        df_obtv = df_obtv_all.iloc[0:0].copy()
    for col in ["VL_PAGO_OBTV_CONV"]:
        if col in df_obtv.columns:
            df_obtv[col] = parse_monetary_series(df_obtv[col])
    df_obtv = df_obtv.loc[:, order_columns_alpha(df_obtv)] if not df_obtv.empty else df_obtv

    # --- montar abas e dicionário
    dfs = {
        SHEET_CONV: df_conv,
        SHEET_PAG:  df_pag,
        SHEET_TRIB: df_trib,
        SHEET_OBTV: df_obtv,
    }
    dict_df = build_dictionary_df(dfs)

    # --- info_execucao
    lm_conv_utc, lm_conv_sp = http_last_modified_to_strings(lm_conv)
    lm_pag_utc,  lm_pag_sp  = http_last_modified_to_strings(lm_pag)
    lm_trib_utc, lm_trib_sp = http_last_modified_to_strings(lm_trib)
    lm_obtv_utc, lm_obtv_sp = http_last_modified_to_strings(lm_obtv)

    extras = {
        "last_modified_convenio_utc": lm_conv_utc,
        "last_modified_convenio_sp": lm_conv_sp,
        "last_modified_pagamento_utc": lm_pag_utc,
        "last_modified_pagamento_sp": lm_pag_sp,
        "last_modified_pagamento_tributo_utc": lm_trib_utc,
        "last_modified_pagamento_tributo_sp": lm_trib_sp,
        "last_modified_obtv_convenente_utc": lm_obtv_utc,
        "last_modified_obtv_convenente_sp": lm_obtv_sp,
    }
    stats = {
        "pagamento": len(df_pag),
        "pagamento_tributo": len(df_trib),
        "obtv_convenente": len(df_obtv),
    }
    info_df = build_info_sheet_df(
        run_dt_local, data_carga_raw, conv_encontrados, len(ids_lista), extras, stats
    )

    # pastas por data de EXECUÇÃO
    y, m, d = run_dt_local.strftime("%Y"), run_dt_local.strftime("%m"), run_dt_local.strftime("%d")
    dir_backup = DIR_OUT / "backups" / y / m / d
    dir_current = DIR_OUT / "current"
    dir_backup.mkdir(parents=True, exist_ok=True)
    dir_current.mkdir(parents=True, exist_ok=True)

    stamp = run_dt_local.strftime("%Y%m%d_%H%M%S")
    fname = f"siconv_convenio_filtrado_{stamp}.xlsx"
    path_xlsx_backup  = dir_backup / fname
    path_xlsx_current = dir_current / "siconv_convenio_filtrado_atual.xlsx"

    write_excel_with_formats(dfs, dict_df, info_df, path_xlsx_backup)

    try:
        shutil.copy2(path_xlsx_backup, path_xlsx_current)
        log(f"📤 Atualizado current: {path_xlsx_current}")
    except PermissionError:
        log("⚠️  Arquivo atual em uso (Excel ou OneDrive). Cópia ignorada.")

    log(f"📁 Arquivo salvo em: {path_xlsx_backup}")
    log("✅ Execução de teste concluída (sem state).")
    return 0

if __name__ == "__main__":
    sys.exit(main())
# ============================================================
